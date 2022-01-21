import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

ktt_orderfile_name = ('快团团订单.xlsx')
workbook = openpyxl.load_workbook(ktt_orderfile_name)

source_sheet = workbook['商品列表']
order_sheet = workbook.create_sheet('订单信息')

r = source_sheet.max_row + 1
print('共计' + str(r-2) + '条数据\n')
for r in range(1,r):
    order_sheet.cell(row=r, column=1).value = source_sheet.cell(row=r, column=1).value
    order_sheet.cell(row=r, column=2).value = source_sheet.cell(row=r, column=24).value
    order_sheet.cell(row=r, column=3).value = str(source_sheet.cell(row=r, column=6).value)
    order_sheet.cell(row=r, column=4).value = source_sheet.cell(row=r, column=30).value
    order_sheet.cell(row=r, column=5).value = source_sheet.cell(row=r, column=31).value
    order_sheet.cell(row=r, column=6).value = source_sheet.cell(row=r, column=35).value
    order_sheet.cell(row=r, column=7).value = source_sheet.cell(row=r, column=9).value
    order_sheet.cell(row=r, column=8).value = source_sheet.cell(row=r, column=8).value
    order_sheet.cell(row=r, column=9).value = source_sheet.cell(row=r, column=36).value
    order_sheet.cell(row=r, column=10).value = source_sheet.cell(row=r, column=16).value
    order_sheet.cell(row=r, column=11).value = source_sheet.cell(row=r, column=13).value
    print('正在写入第' + str(r) + '行数据')

for i in range(2,r+1):
    order_sheet.cell(row=i, column=3).value = order_sheet.cell(row=i, column=3).value.split(' ')[0]
print('数据写入完成！\n')
workbook.save(ktt_orderfile_name)
input('按任意键结束……')